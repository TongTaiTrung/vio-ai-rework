from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
import matplotlib.pyplot as plt
import matplotlib
from PIL import Image
import io, base64, os, tempfile
from io import BytesIO

matplotlib.rcParams["font.family"] = "Times New Roman"
DEBUG = False

joint_title = {
    "left_elbow": "Khủy tay trái",
    "right_elbow": "Khủy tay phải",
    "left_shoulder": "Vai trái",
    "right_shoulder": "Vai phải",
    "left_knee": "Chân trái",
    "right_knee": "Chân phải",
    "left_hip": "Hông trái",
    "right_hip": "Hông phải",
}

def export_excel(joint_summary, score, studentCode):
    wb = Workbook()
    ws = wb.active
    ws.title = "So sánh góc"

    headers = ["Khớp :"] + [joint_title[j] for j in joint_summary.keys()]
    ws.append(headers)

    row_student = ["Học Sinh"] + [f"{vals['student_avg']:.1f}˚" for vals in joint_summary.values()]
    ws.append(row_student)

    row_sample = ["Mẫu"] + [f"{vals['sample_avg']:.1f}˚" for vals in joint_summary.values()]
    ws.append(row_sample)

    comments = ["Đánh giá học sinh"]
    for joint, vals in joint_summary.items():
        diff = vals["student_avg"] - vals["sample_avg"]
        if abs(diff) < 5:
            comments.append("Tương đồng với mẫu")
        elif diff > 0:
            comments.append(f"Rộng hơn mẫu {abs(round(diff, 1))}˚")
        else:
            comments.append(f"Hẹp hơn mẫu {abs(round(diff, 1))}˚")
    ws.append(comments)
    
    ws.merge_cells("B5:I5")
    ws.merge_cells("B6:I6")
    lastRows = [["A5","Điểm tổng"], ["A6","Mã số học sinh"],
                ["B5", f"{score:.1f}/100.0"], ["B6", studentCode]]
    for lastRow in lastRows:
        ws[lastRow[0]] = lastRow[1]
        ws[lastRow[0]].font = Font(name="Times New Roman", size=12, bold=True, italic=True)
        ws[lastRow[0]].alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(name="Times New Roman", size=10)
            cell.border = thin_border

    for cell in ws[1]:
        cell.font = Font(name="Times New Roman", size=12, bold=True, italic=True)

    for row_idx in range(2, ws.max_row+1):
        ws[f"A{row_idx}"].font = Font(name="Times New Roman", size=12, bold=True, italic=True)

    joints = [joint_title[j] for j in joint_summary.keys()]
    student_vals = [vals['student_avg'] for vals in joint_summary.values()]
    sample_vals = [vals['sample_avg'] for vals in joint_summary.values()]

    x = range(len(joints))
    plt.figure(figsize=(8, 5))
    bars1 = plt.bar([i - 0.2 for i in x], student_vals, width=0.3,
                    label="Học Sinh", color="#FF6600")
    bars2 = plt.bar([i + 0.2 for i in x], sample_vals, width=0.3,
                    label="Mẫu", color="#0066FF")

    plt.xticks(x, joints, rotation=30, ha="right")
    plt.ylabel("Góc (độ)")
    plt.title("So sánh góc giữa Học sinh và Mẫu")
    plt.legend()
    plt.tight_layout()

    for bar in bars1 + bars2:
        height = bar.get_height()
        plt.text(bar.get_x() + bar.get_width()/2, height + 0.5,
                 f"{height:.1f}˚", ha="center", va="bottom",
                 fontsize=9, fontname="Times New Roman")

    img_buffer = io.BytesIO()
    plt.savefig(img_buffer, format="png", dpi=225)
    plt.close()
    img_buffer.seek(0)
    chart_bytes = img_buffer.read()

    with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp:
        tmp.write(chart_bytes)
        tmp_path = tmp.name

    img = XLImage(tmp_path)
    ws.add_image(img, "A10")

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    excel_base64 = base64.b64encode(excel_buffer.read()).decode("utf-8")

    if DEBUG:
        os.makedirs("reports", exist_ok=True)
        with open(os.path.join("reports", f"{studentCode}_chart.png"), "wb") as f:
            f.write(chart_bytes)
        wb.save(os.path.join("reports", f"{studentCode}.xlsx"))
        
    chart_image = Image.open(BytesIO(chart_bytes))
    return {
        "excel_bytes": base64.b64decode(excel_base64),
        "chart_image": chart_image
    }